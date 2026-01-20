"""
Excel Exporter for VF-7 Land Records
Exports structured data to Excel with multiple sheets
"""

import json
import os
from datetime import datetime
from typing import List, Dict
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[WARN] openpyxl not installed. Run: pip install openpyxl")

logger = logging.getLogger(__name__)


class VF7ExcelExporter:
    """Export VF-7 data to Excel with multiple sheets"""
    
    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        
        # Styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_single_sheet_workbook(self, results: List[Dict], output_path: str,
                                     district_name: str = "", taluka_name: str = "") -> str:
        """
        Create Excel with ONE sheet containing all data (flattened)
        
        Args:
            results: List of structured VF-7 results
            output_path: Path to save Excel file
            district_name: District name for metadata
            taluka_name: Taluka name for metadata
        
        Returns:
            Path to created Excel file
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Data"
        
        # Headers - ALL fields in one row
        headers = [
            # Location
            "Village Code", "Village Name", "District", "Taluka",
            # Property Identity
            "Survey Number", "UPIN", "Khata Number", "Old Survey Number",
            # Land Details
            "Total Area (Raw)", "Area (Hectare)", "Area (Are)", "Area (SqM)", "Area (Sq Yd)",
            "Assessment Tax", "Tenure", "Land Use", "Farm Name",
            # Owners (comma-separated)
            "Owner Names", "Owner Count",
            # Encumbrances (comma-separated)
            "Encumbrances", "Encumbrance Count",
            # Other
            "Data Status Time", "Remarks"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        # Data rows
        row = 2
        for result in results:
            if not result.get('success', True):  # Skip failed if marked
                continue
            
            location = result.get('location', {})
            prop_id = result.get('property_identity', {})
            land = result.get('land_details', {})
            owners = result.get('owners', [])
            encumbrances = result.get('rights_and_remarks', {}).get('entry_details', [])
            meta = result.get('meta', {})
            
            # Flatten owners
            owner_names = ", ".join([o.get('owner_name', '') for o in owners])
            
            # Flatten encumbrances
            enc_descriptions = "; ".join([
                f"{e.get('type', '')}: {e.get('description', '')[:50]}" 
                for e in encumbrances
            ])
            
            data = [
                result.get('village_code', ''),
                location.get('village', {}).get('name_local', ''),
                location.get('district', {}).get('name_local', ''),
                location.get('taluka', {}).get('name_local', ''),
                prop_id.get('survey_number', ''),
                prop_id.get('upin', ''),
                prop_id.get('khata_number', ''),
                prop_id.get('old_survey_number', ''),
                land.get('total_area_raw', ''),
                land.get('area_hectare', ''),
                land.get('area_are', ''),
                land.get('area_sqm', ''),
                land.get('area_sq_yd', ''),
                land.get('assessment_tax', ''),
                land.get('tenure_local', '') or land.get('tenure', ''),
                land.get('land_use_local', '') or land.get('land_use', ''),
                land.get('farm_name', ''),
                owner_names,
                len(owners),
                enc_descriptions,
                len(encumbrances),
                meta.get('data_status_time_local', ''),
                land.get('remarks', '')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if col <= 4:  # Key columns
                    cell.font = Font(bold=True)
            
            row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save
        wb.save(output_path)
        logger.info(f"Single-sheet Excel exported: {output_path}")
        
        return output_path
    
    def create_workbook(self, results: List[Dict], output_path: str, 
                       district_name: str = "", taluka_name: str = "") -> str:
        """
        Create Excel workbook with multiple sheets
        
        Args:
            results: List of structured VF-7 results
            output_path: Path to save Excel file
            district_name: District name for metadata
            taluka_name: Taluka name for metadata
        
        Returns:
            Path to created Excel file
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self._create_summary_sheet(wb, results, district_name, taluka_name)
        self._create_properties_sheet(wb, results)
        self._create_owners_sheet(wb, results)
        self._create_encumbrances_sheet(wb, results)
        
        # Save
        wb.save(output_path)
        logger.info(f"Excel exported: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, wb, results: List[Dict], district: str, taluka: str):
        """Create summary sheet with statistics"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "VF-7 Land Records Export Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        metadata = [
            ("Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("District:", district),
            ("Taluka:", taluka),
            ("Total Records:", len(results)),
            ("Successful:", sum(1 for r in results if r.get('success', False))),
            ("Failed:", sum(1 for r in results if not r.get('success', False))),
        ]
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Statistics
        row += 2
        ws[f'A{row}'] = "Statistics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        successful_results = [r for r in results if r.get('success', False)]
        
        if successful_results:
            total_owners = sum(len(r.get('owners', [])) for r in successful_results)
            total_encumbrances = sum(
                len(r.get('rights_and_remarks', {}).get('entry_details', [])) 
                for r in successful_results
            )
            
            stats = [
                ("Total Owners:", total_owners),
                ("Total Encumbrances:", total_encumbrances),
                ("Avg Owners per Property:", round(total_owners / len(successful_results), 2)),
                ("Avg Encumbrances per Property:", round(total_encumbrances / len(successful_results), 2)),
            ]
            
            for label, value in stats:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def _create_properties_sheet(self, wb, results: List[Dict]):
        """Create properties sheet with all land details"""
        ws = wb.create_sheet("Properties")
        
        # Headers
        headers = [
            "Village Code", "Village Name", "Survey Number", "UPIN",
            "Khata Number", "Old Survey Number",
            "Total Area (Raw)", "Area (Hectare)", "Area (Are)", "Area (SqM)", "Area (Sq Yd)",
            "Assessment Tax", "Tenure", "Land Use", "Farm Name",
            "Owner Count", "Encumbrance Count",
            "Data Status Time", "Remarks"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        row = 2
        for result in results:
            if not result.get('success', False):
                continue
            
            location = result.get('location', {})
            prop_id = result.get('property_identity', {})
            land = result.get('land_details', {})
            meta = result.get('meta', {})
            
            data = [
                result.get('village_code', ''),
                location.get('village', {}).get('name_local', ''),
                prop_id.get('survey_number', ''),
                prop_id.get('upin', ''),
                prop_id.get('khata_number', ''),
                prop_id.get('old_survey_number', ''),
                land.get('total_area_raw', ''),
                land.get('area_hectare', ''),
                land.get('area_are', ''),
                land.get('area_sqm', ''),
                land.get('area_sq_yd', ''),
                land.get('assessment_tax', ''),
                land.get('tenure_local', '') or land.get('tenure', ''),
                land.get('land_use_local', '') or land.get('land_use', ''),
                land.get('farm_name', ''),
                len(result.get('owners', [])),
                len(result.get('rights_and_remarks', {}).get('entry_details', [])),
                meta.get('data_status_time_local', ''),
                land.get('remarks', '')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col <= 4:  # Key columns
                    cell.font = Font(bold=True)
            
            row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_owners_sheet(self, wb, results: List[Dict]):
        """Create owners sheet with all ownership details"""
        ws = wb.create_sheet("Owners")
        
        # Headers
        headers = [
            "Village Code", "Village Name", "Survey Number", "UPIN",
            "Owner Name", "Father/Husband Name", "Ownership Type",
            "Entry Number", "Share", "Address/Remark"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        row = 2
        for result in results:
            if not result.get('success', False):
                continue
            
            location = result.get('location', {})
            prop_id = result.get('property_identity', {})
            village_code = result.get('village_code', '')
            village_name = location.get('village', {}).get('name_local', '')
            survey = prop_id.get('survey_number', '')
            upin = prop_id.get('upin', '')
            
            owners = result.get('owners', [])
            
            for owner in owners:
                data = [
                    village_code,
                    village_name,
                    survey,
                    upin,
                    owner.get('owner_name', ''),
                    owner.get('owner_father_or_husband_name', ''),
                    owner.get('ownership_type', ''),
                    owner.get('entry_number', ''),
                    owner.get('share', ''),
                    owner.get('owner_address_or_remark', '')
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                
                row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_encumbrances_sheet(self, wb, results: List[Dict]):
        """Create encumbrances sheet with all rights and remarks"""
        ws = wb.create_sheet("Encumbrances")
        
        # Headers
        headers = [
            "Village Code", "Village Name", "Survey Number", "UPIN",
            "Entry Number", "Entry Date", "Type", "Description",
            "Amount", "Status", "Office", "Order Reference"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data rows
        row = 2
        for result in results:
            if not result.get('success', False):
                continue
            
            location = result.get('location', {})
            prop_id = result.get('property_identity', {})
            village_code = result.get('village_code', '')
            village_name = location.get('village', {}).get('name_local', '')
            survey = prop_id.get('survey_number', '')
            upin = prop_id.get('upin', '')
            
            encumbrances = result.get('rights_and_remarks', {}).get('entry_details', [])
            
            for enc in encumbrances:
                data = [
                    village_code,
                    village_name,
                    survey,
                    upin,
                    enc.get('entry_no', ''),
                    enc.get('entry_date', ''),
                    enc.get('type', ''),
                    enc.get('description', ''),
                    enc.get('amount', ''),
                    enc.get('status', ''),
                    enc.get('office', ''),
                    enc.get('order_reference', '')
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                
                row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            width = 20 if col <= 4 else 25
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def export_from_json_files(self, json_files: List[str], output_path: str,
                               district_name: str = "", taluka_name: str = "") -> str:
        """
        Export multiple JSON files to single Excel
        
        Args:
            json_files: List of JSON file paths
            output_path: Output Excel path
            district_name: District name
            taluka_name: Taluka name
        
        Returns:
            Path to created Excel file
        """
        all_results = []
        
        for json_file in json_files:
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                    # Handle different JSON formats
                    if 'results' in data:
                        results = data['results']
                    elif isinstance(data, list):
                        results = data
                    else:
                        results = [data]
                    
                    # Extract structured data if needed
                    for result in results:
                        if 'structured' in result:
                            all_results.append(result['structured'])
                        else:
                            all_results.append(result)
            
            except Exception as e:
                logger.error(f"Error loading {json_file}: {e}")
                continue
        
        return self.create_workbook(all_results, output_path, district_name, taluka_name)
    
    def export_from_directory(self, input_dir: str, output_path: str = None,
                             district_name: str = "", taluka_name: str = "") -> str:
        """
        Export all JSON files from directory to Excel
        
        Args:
            input_dir: Directory containing JSON files
            output_path: Output Excel path (auto-generated if None)
            district_name: District name
            taluka_name: Taluka name
        
        Returns:
            Path to created Excel file
        """
        # Find all JSON files
        json_files = []
        for root, dirs, files in os.walk(input_dir):
            for file in files:
                if file.endswith('.json') and not file.startswith('global_summary'):
                    json_files.append(os.path.join(root, file))
        
        if not json_files:
            raise ValueError(f"No JSON files found in {input_dir}")
        
        logger.info(f"Found {len(json_files)} JSON files")
        
        # Auto-generate output path
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(input_dir, f"vf7_export_{timestamp}.xlsx")
        
        return self.export_from_json_files(json_files, output_path, district_name, taluka_name)


def main():
    """Example usage"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python excel_exporter.py <input_dir> [output.xlsx]")
        print("  python excel_exporter.py output/")
        print("  python excel_exporter.py output/ results.xlsx")
        return
    
    input_dir = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    exporter = VF7ExcelExporter()
    
    try:
        result_path = exporter.export_from_directory(input_dir, output_path)
        print(f"\n✅ Excel exported: {result_path}")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
